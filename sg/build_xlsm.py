# -*- coding: utf-8 -*-
"""근무표 자동화 템플릿 빌더 (집계 수식 + 조건부서식 + 검증 시트)"""
import openpyxl, calendar, datetime, sys
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter as L

YEAR = int(sys.argv[1]) if len(sys.argv) > 1 else 2026
MONTH = int(sys.argv[2]) if len(sys.argv) > 2 else 6
OUTDIR = sys.argv[3] if len(sys.argv) > 3 else "."
OUT = f"{OUTDIR}/근무표_{YEAR}년{MONTH:02d}월.xlsx"

ndays = calendar.monthrange(YEAR, MONTH)[1]
WD = ['월','화','수','목','금','토','일']

FILL_HEADER = PatternFill('solid', fgColor='4472C4')
FILL_SAT    = PatternFill('solid', fgColor='DDEBF7')
FILL_SUN    = PatternFill('solid', fgColor='FCE4EC')
FILL_REST   = PatternFill('solid', fgColor='F2F2F2')
FILL_LEAVE  = PatternFill('solid', fgColor='FFF2CC')
FILL_DAEHYU = PatternFill('solid', fgColor='E2EFDA')
thin = Side(style='thin', color='BFBFBF')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
F_BASE = Font(name='맑은 고딕', size=10)
F_HEAD = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

employees = ['센터장','김영지','문세동','정은상','이병호','현병택','운전원','보조원',
             '김경희','김재덕','김민재','이영화','김선미']
NO_COUNT = {'운전원','보조원'}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "근무편성표"

ws.cell(1,1,'구분')
for d in range(1, ndays+1):
    ws.cell(1, 1+d, d)
col = 1+ndays
for i,h in enumerate(['근무일','휴무','대휴','연차','근무시간','비고']):
    ws.cell(1, col+1+i, h)
ws.cell(2,1,'요일')
for d in range(1, ndays+1):
    ws.cell(2, 1+d, WD[datetime.date(YEAR,MONTH,d).weekday()])

start_row = 3
for ei, name in enumerate(employees):
    ws.cell(start_row+ei, 1, name)
last_data_row = start_row + len(employees) - 1
day_first, day_last = 2, 1+ndays
agg = col+1
C_WORK, C_REST, C_DAEHYU, C_LEAVE, C_HOURS, C_NOTE = agg, agg+1, agg+2, agg+3, agg+4, agg+5
fL, lL = L(day_first), L(day_last)

for ei, name in enumerate(employees):
    r = start_row+ei
    rng = f"{fL}{r}:{lL}{r}"
    ws.cell(r, C_REST,   f'=COUNTIF({rng},"휴")+COUNTIF({rng},"*(휴)*")')
    ws.cell(r, C_DAEHYU, f'=COUNTIF({rng},"대휴")')
    ws.cell(r, C_LEAVE,  f'=COUNTIF({rng},"연차")+COUNTIF({rng},"*반차*")+COUNTIF({rng},"교육")')
    ws.cell(r, C_WORK,   f'=COUNTA({rng})-{L(C_REST)}{r}-{L(C_DAEHYU)}{r}-{L(C_LEAVE)}{r}')
    ws.cell(r, C_HOURS, '기준제외' if name in NO_COUNT else f'={L(C_WORK)}{r}*8')

for c in range(1, C_NOTE+1):
    h=ws.cell(1,c); h.fill=FILL_HEADER; h.font=F_HEAD; h.alignment=CENTER; h.border=BORDER
    d2=ws.cell(2,c); d2.font=F_BASE; d2.alignment=CENTER; d2.border=BORDER
for d in range(1, ndays+1):
    wd=datetime.date(YEAR,MONTH,d).weekday(); c=1+d
    if wd==5: ws.cell(2,c).fill=FILL_SAT
    elif wd==6: ws.cell(2,c).fill=FILL_SUN
for ei in range(len(employees)):
    r=start_row+ei
    for c in range(1, C_NOTE+1):
        cell=ws.cell(r,c); cell.font=F_BASE; cell.alignment=CENTER; cell.border=BORDER
    for d in range(1, ndays+1):
        wd=datetime.date(YEAR,MONTH,d).weekday(); c=1+d
        if wd==5: ws.cell(r,c).fill=FILL_SAT
        elif wd==6: ws.cell(r,c).fill=FILL_SUN

ws.column_dimensions['A'].width=10
for d in range(1, ndays+1): ws.column_dimensions[L(1+d)].width=8
for c in range(agg, C_NOTE): ws.column_dimensions[L(c)].width=9
ws.column_dimensions[L(C_NOTE)].width=16
for r in range(1, last_data_row+1): ws.row_dimensions[r].height=22
ws.freeze_panes='B3'

data_rng=f"{fL}{start_row}:{lL}{last_data_row}"
ws.conditional_formatting.add(data_rng, CellIsRule(operator='equal', formula=['"휴"'], fill=FILL_REST))
ws.conditional_formatting.add(data_rng, CellIsRule(operator='equal', formula=['"대휴"'], fill=FILL_DAEHYU))
ws.conditional_formatting.add(data_rng, CellIsRule(operator='equal', formula=['"연차"'], fill=FILL_LEAVE))
ws.conditional_formatting.add(data_rng, CellIsRule(operator='equal', formula=['"교육"'], fill=FILL_LEAVE))
ws.conditional_formatting.add(data_rng, FormulaRule(formula=[f'ISNUMBER(SEARCH("반차",{fL}{start_row}))'], fill=FILL_LEAVE))

vs=wb.create_sheet("검증")
for i,t in enumerate(['검증항목','직원','결과']):
    c=vs.cell(1,i+1,t); c.font=F_HEAD; c.fill=FILL_HEADER; c.alignment=CENTER
vr=2
for ei,name in enumerate(employees):
    if name in NO_COUNT: continue
    r=start_row+ei
    vs.cell(vr,1,'휴무일수')
    vs.cell(vr,2,name)
    vs.cell(vr,3, f'=IF(OR(근무편성표!{L(C_REST)}{r}<8,근무편성표!{L(C_REST)}{r}>10),"⚠ 휴무 "&근무편성표!{L(C_REST)}{r}&"일 (8~10 권장)","OK "&근무편성표!{L(C_REST)}{r}&"일")')
    vr+=1
# 빈칸(미입력) 검증
vs.cell(vr,1,'미입력 칸')
vs.cell(vr,2,'전체')
total_cells=len(employees)*ndays
vs.cell(vr,3, f'=IF(COUNTBLANK(근무편성표!{fL}{start_row}:{lL}{last_data_row})>0,"⚠ 빈칸 "&COUNTBLANK(근무편성표!{fL}{start_row}:{lL}{last_data_row})&"개","OK 모두 입력됨")')
vs.column_dimensions['A'].width=14
vs.column_dimensions['B'].width=10
vs.column_dimensions['C'].width=34

wb.save(OUT)
print("SAVED", OUT, "days=", ndays)
